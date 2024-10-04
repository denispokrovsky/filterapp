import streamlit as st
import pandas as pd
import re
from io import BytesIO
from rapidfuzz import fuzz
from openpyxl import load_workbook
import openai
from langchain.prompts import PromptTemplate

# Streamlit app layout
st.set_page_config(page_title="::: мониторинг новостного потока :::", layout="wide")

st.title('Фильтр новостного файла в формате СКАН-Интерфакс на релевантность и значимость!')
st.write("Загружайте и выгружайте!")

# File uploader
uploaded_file = st.file_uploader("Выбери Excel файл", type=["xlsx"])

# Access the OpenAI API key from Streamlit secrets
openai_api_key = st.secrets["OPENAI_API_KEY2"]
openai.api_key = openai_api_key  # Set OpenAI API key

# Define prompt templates for LangChain
risk_prompt_template = PromptTemplate(
    input_variables=["text", "company"],
    template="Текст: {text}\nКомпания: {company}\nЕсть ли риск убытка для этой компании в ближайшие шесть месяцев? Ответьте 'Риск убытка' или 'Нет риска убытка'."
)

comment_prompt_template = PromptTemplate(
    input_variables=["text", "company"],
    template="Текст: {text}\nКомпания: {company}\nСколько примерно может потерять компания? Дайте комментарий не более 200 слов."
)

# Function to call OpenAI's API with the new ChatCompletion method
def call_openai(prompt):
    response = openai.ChatCompletion.create(
        model="gpt-4o-mini",  # Updated model to gpt-4o-mini
        messages=[
            {"role": "system", "content": "You are a financial credit analyst. You assess probability of short-term credit risk of the company or a bank."},
            {"role": "user", "content": prompt}
        ]
    )
    return response['choices'][0]['message']['content']

# Function to process the Excel file and add new columns without LLM
def process_excel_without_llm(file, sample_file, similarity_threshold=65):
    # Load all sheets from the uploaded Excel file
    excel_file = pd.ExcelFile(file)
    sheets = {sheet_name: excel_file.parse(sheet_name) for sheet_name in excel_file.sheet_names}

    # Access the required sheet 'Публикации'
    df = sheets['Публикации'].copy()

    # Track original number of news (X)
    original_news_count = len(df)

    # Step 1: Generate the list of unique companies from the 'Объект' column
    unique_companies = df['Объект'].dropna().unique().tolist()

    # Step 2: Fuzzy filter out similar news for the same company/bank
    def fuzzy_deduplicate(df, column, threshold=65):
        seen_texts = []
        indices_to_keep = []
        for i, text in enumerate(df[column]):
            if pd.isna(text):
                indices_to_keep.append(i)
                continue
            text = str(text)
            if not seen_texts or all(fuzz.ratio(text, seen) < threshold for seen in seen_texts):
                seen_texts.append(text)
                indices_to_keep.append(i)
        return df.iloc[indices_to_keep]

    # Apply fuzzy deduplication on 'Выдержки из текста' column
    df_deduplicated = df.groupby('Объект').apply(lambda x: fuzzy_deduplicate(x, 'Выдержки из текста', similarity_threshold)).reset_index(drop=True)

    # Track the number of remaining news (Z)
    remaining_news_count = len(df_deduplicated)

    # Calculate number of duplicates removed (Y)
    duplicates_removed = original_news_count - remaining_news_count

    # Step 3: Define relevance assessment (including Russian-specific keywords)
    direct_keywords = ['убыток', 'прибыль', 'судебное дело', 'банкротство', 'потеря', 'миллиард', 'млн', 'миллиардов', 'миллионов', 'выручка']
    indirect_keywords = ['аналитик', 'комментарий', 'прогноз', 'отчет', 'заявление']

    def assess_relevance(text, company):
        if pd.isna(text):
            return 'н/д'
        text = text.lower()
        direct_relevance = any(keyword in text for keyword in direct_keywords) and company.lower() in text
        indirect_relevance = any(keyword in text for keyword in indirect_keywords)
        if direct_relevance and not indirect_relevance:
            return 'материальна'
        elif indirect_relevance:
            return 'нематериальная'
        else:
            return 'н/д'

    df_deduplicated['Relevance'] = df_deduplicated.apply(lambda row: assess_relevance(row['Выдержки из текста'], row['Объект']), axis=1)

    # Step 4: Sentiment assessment
    negative_keywords = ['убыток', 'потеря', 'снижение', 'упадок', 'падение']
    positive_keywords = ['прибыль', 'рост', 'увеличение', 'подъем']

    def assess_sentiment(text):
        if pd.isna(text):
            return 'нейтрально'
        text = text.lower()
        if any(word in text for word in negative_keywords):
            return 'негатив'
        elif any(word in text for word in positive_keywords):
            return 'позитив'
        else:
            return 'нейтрально'

    df_deduplicated['Sentiment'] = df_deduplicated['Выдержки из текста'].apply(assess_sentiment)

    # Step 5: Materiality assessment
    def assess_probable_materiality(text):
        if pd.isna(text):
            return 'н/д'
        match = re.search(r'(\d+)\s*млрд\s*руб', text.lower())
        if match:
            return 'значительна'
        elif 'миллион' in text.lower():
            return 'значительна'
        else:
            return 'незначительна'

    df_deduplicated['Materiality_Level'] = df_deduplicated['Выдержки из текста'].apply(assess_probable_materiality)

    # Step 6: Prepare summary for "Сводка" sheet
    dashboard_summary = df_deduplicated.groupby('Объект').agg(
        News_Count=('Выдержки из текста', 'count'),
        Significant_Texts=('Materiality_Level', lambda x: (x == 'значительна').sum()),
        Negative_Texts=('Sentiment', lambda x: (x == 'негатив').sum()),
        Positive_Texts=('Sentiment', lambda x: (x == 'позитив').sum()),
        Risk_Level=('Materiality_Level', lambda x: 'высокий' if 'значительна' in x.values else 'низкий')
    ).reset_index()

    # Sort the summary by News_Count first and Significant_Texts second (both in descending order)
    dashboard_summary_sorted = dashboard_summary.sort_values(by=['Significant_Texts', 'News_Count'], ascending=[False, False])

    # Save the final file to a BytesIO buffer
    output = BytesIO()
    book = load_workbook(sample_file)

    # Rename columns for display in Streamlit
    dashboard_summary_sorted.columns = [
        'Компания',
        'Всего публикаций',
        'Из них: материальных',
        'Из них: негативных',
        'Из них: позитивных',
        'Уровень материального негатива'
    ]

    filtered_news = df_deduplicated[df_deduplicated['Relevance'] == 'материальна']
    filtered_news = filtered_news.drop_duplicates(subset=['Объект', 'Выдержки из текста']).reset_index(drop=True)

    # Write sorted data to the "Сводка" sheet
    dashboard_sheet = book['Сводка']
    for idx, row in dashboard_summary_sorted.iterrows():
        dashboard_sheet[f'E{4 + idx}'] = row['Компания']
        dashboard_sheet[f'F{4 + idx}'] = row['Всего публикаций']
        dashboard_sheet[f'G{4 + idx}'] = row['Из них: материальных']
        dashboard_sheet[f'H{4 + idx}'] = row['Из них: негативных']
        dashboard_sheet[f'I{4 + idx}'] = row['Из них: позитивных']
        dashboard_sheet[f'J{4 + idx}'] = row['Уровень материального негатива']

    # Save the final file to a BytesIO buffer
    output = BytesIO()
    book.save(output)
    output.seek(0)

    return output, df_deduplicated, original_news_count, duplicates_removed, remaining_news_count, dashboard_summary_sorted

# Function to apply LLM analysis and generate a new Excel file
def apply_llm_analysis(df_deduplicated, previous_file):
    # Load the previously generated file (processed without LLM)
    book = load_workbook(previous_file)

    # Apply LLM for Risk and Comment assessment
    df_deduplicated['Risk of loss'] = df_deduplicated.apply(lambda row: call_openai(risk_prompt_template.format(text=row['Выдержки из текста'], company=row['Объект'])), axis=1)
    df_deduplicated['Comment'] = df_deduplicated.apply(lambda row: call_openai(comment_prompt_template.format(text=row['Выдержки из текста'], company=row['Объект'])), axis=1)

    # Filter rows where 'Risk of loss' is 'Риск убытка'
    filtered_llm_news = df_deduplicated[df_deduplicated['Risk of loss'] == 'Риск убытка']

    # Write filtered LLM results to the existing "Анализ" sheet in the Excel file
    analysis_sheet = book['Анализ']

    for idx, row in filtered_llm_news.iterrows():
        analysis_sheet[f'E{4 + idx}'] = row['Объект']
        analysis_sheet[f'F{4 + idx}'] = row['Заголовок']
        analysis_sheet[f'G{4 + idx}'] = row['Risk of loss']
        analysis_sheet[f'H{4 + idx}'] = row['Comment']
        analysis_sheet[f'I{4 + idx}'] = row['Выдержки из текста']

    # Save the final file with LLM results to a BytesIO buffer
    output_llm = BytesIO()
    book.save(output_llm)
    output_llm.seek(0)

    return output_llm, filtered_llm_news

# Handle file upload and processing
if uploaded_file is not None:
    # Store the path to the sample Excel file for formatting
    sample_file = "sample_file.xlsx"

    # Step 1: Process the file without LLM analysis and display the dashboard summary
    processed_file, filtered_table, original_news_count, duplicates_removed, remaining_news_count, dashboard_summary_sorted = process_excel_without_llm(uploaded_file, sample_file)

    # Save the processed file locally as `processed_news_without_llm.xlsx`
    with open("/tmp/processed_news_without_llm.xlsx", "wb") as f:
        f.write(processed_file.getbuffer())

    # Display the filtered news as it appears in Excel
    st.write(f"Из {original_news_count} новостных сообщений удалены {duplicates_removed} дублирующих. Осталось {remaining_news_count}.")
    
    st.write("Только материальные новости:")
    st.dataframe(filtered_table[['Объект', 'Relevance', 'Sentiment', 'Materiality_Level', 'Заголовок', 'Выдержки из текста']])

    # Display the initial dashboard summary without LLM
    st.write("Сводка без анализа LLM:")
    st.dataframe(dashboard_summary_sorted)

    # Provide a download button for the first processed file (without LLM)
    st.download_button(
        label="Скачать обработанный файл (без LLM)",
        data=processed_file,
        file_name="processed_news_without_llm.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Step 2: Introduce a button to proceed with LLM analysis
    if st.button("Применить анализ LLM"):
        # Apply LLM analysis and generate a new processed file with LLM results
        processed_llm_file, new_dashboard_summary = apply_llm_analysis(filtered_table, "/tmp/processed_news_without_llm.xlsx")

        # Display the new dashboard summary with LLM analysis
        st.write("Сводка с анализом LLM (Риск убытка):")
        st.dataframe(new_dashboard_summary)

        # Provide a download button for the second processed file (with LLM)
        st.download_button(
            label="Скачать обработанный файл (с LLM)",
            data=processed_llm_file,
            file_name="processed_news_with_llm.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
